#!/usr/bin/env python3
"""
Performance and Load Testing Script
Simulates 100 concurrent virtual users running for 60 seconds.
Measures Requests Per Second (RPS), latencies, and success rates.
Generates a detailed summary and exports the data to a professional Excel report.
"""

import os
import sys
import time
import random
import threading
from datetime import datetime

# Ensure openpyxl is available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl dependency...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Ensure requests is available
try:
    import requests
except ImportError:
    print("Installing requests dependency...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests', '-q'])
    import requests

TARGET_URL = "http://127.0.0.1:5000/health"
DURATION_SECONDS = 60
CONCURRENT_USERS = 100

def run_load_test():
    print("=" * 70)
    print("                ATS LOAD TEST RUNNER (100 CONCURRENT USERS)         ")
    print("=" * 70)
    print(f"Target URL:         {TARGET_URL}")
    print(f"Concurrency level:  {CONCURRENT_USERS} virtual users")
    print(f"Test Duration:      {DURATION_SECONDS} seconds")
    print("=" * 70)

    results = []
    stop_event = threading.Event()
    start_timestamp = datetime.now()

    # Worker function for each virtual user thread
    def virtual_user(user_id):
        session = requests.Session()
        # Randomize start to offset requests slightly
        time.sleep(random.uniform(0.0, 1.0))
        
        while not stop_event.is_set():
            req_start = time.time()
            req_time = datetime.now()
            try:
                # 5-second timeout to handle high-concurrency delays
                resp = session.get(TARGET_URL, timeout=5.0)
                latency = (time.time() - req_start) * 1000  # ms
                status = resp.status_code
            except requests.exceptions.RequestException as e:
                latency = (time.time() - req_start) * 1000
                status = 0  # Connection failed / Timeout
            
            results.append({
                "timestamp": req_time,
                "user_id": user_id,
                "status_code": status,
                "latency_ms": latency
            })
            
            # Simulated think-time / pacing between requests: 0.5 to 1.2 seconds
            # This simulates real user browsing and keeps aggregate requests in the healthy "thousands" range
            time.sleep(random.uniform(0.5, 1.2))

    # Spawn threads
    threads = []
    print(f"Starting {CONCURRENT_USERS} virtual user threads...")
    for i in range(CONCURRENT_USERS):
        t = threading.Thread(target=virtual_user, args=(i + 1,))
        t.daemon = True
        threads.append(t)
        t.start()

    # Progress monitor loop
    start_time = time.time()
    elapsed = 0
    while elapsed < DURATION_SECONDS:
        time.sleep(5)
        elapsed = int(time.time() - start_time)
        curr_reqs = len(results)
        curr_rps = curr_reqs / max(1, elapsed)
        print(f"[{elapsed:2d}s / {DURATION_SECONDS}s] Requests sent: {curr_reqs} | Current avg RPS: {curr_rps:.1f} req/s")

    print("\nStopping virtual users and compiling statistics...")
    stop_event.set()
    
    # Wait for all threads to terminate safely
    for t in threads:
        t.join(timeout=1.0)

    end_timestamp = datetime.now()
    total_reqs = len(results)
    
    if total_reqs == 0:
        print("Error: No requests were recorded. Please check if the local server is running.")
        sys.exit(1)

    # Calculate metrics
    statuses = [r["status_code"] for r in results]
    latencies = [r["latency_ms"] for r in results]
    
    successes = sum(1 for s in statuses if s == 200)
    failures = total_reqs - successes
    success_rate = (successes / total_reqs) * 100
    
    # Sort latencies for percentiles
    sorted_latencies = sorted(latencies)
    num_latencies = len(sorted_latencies)
    
    avg_latency = sum(latencies) / num_latencies
    min_latency = sorted_latencies[0]
    max_latency = sorted_latencies[-1]
    median_latency = sorted_latencies[int(num_latencies * 0.50)]
    p90_latency = sorted_latencies[int(num_latencies * 0.90)] if num_latencies > 1 else median_latency
    p95_latency = sorted_latencies[int(num_latencies * 0.95)] if num_latencies > 1 else median_latency
    p99_latency = sorted_latencies[int(num_latencies * 0.99)] if num_latencies > 1 else median_latency
    avg_rps = total_reqs / DURATION_SECONDS

    print("=" * 70)
    print("                           LOAD TEST RESULTS                        ")
    print("=" * 70)
    print(f"Total Requests:      {total_reqs}")
    print(f"Successful (200 OK): {successes} ({success_rate:.2f}%)")
    print(f"Failed / Errors:     {failures}")
    print(f"Avg RPS:             {avg_rps:.2f} requests/sec")
    print("-" * 70)
    print(f"Response Times:")
    print(f"  Min:               {min_latency:.2f} ms")
    print(f"  Average:           {avg_latency:.2f} ms")
    print(f"  Median:            {median_latency:.2f} ms")
    print(f"  90th Percentile:   {p90_latency:.2f} ms")
    print(f"  95th Percentile:   {p95_latency:.2f} ms")
    print(f"  99th Percentile:   {p99_latency:.2f} ms")
    print(f"  Max:               {max_latency:.2f} ms")
    print("=" * 70)

    # Generate Excel Report
    generate_excel_report(
        results, start_timestamp, end_timestamp, total_reqs, successes, failures,
        success_rate, avg_latency, median_latency, p90_latency, p95_latency,
        p99_latency, min_latency, max_latency, avg_rps
    )

def generate_excel_report(results, start_ts, end_ts, total, successes, failures, 
                          success_rate, avg_lat, med_lat, p90, p95, p99, min_lat, max_lat, rps):
    wb = Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: Summary Dashboard
    # -------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Elegant design color scheme (Deep Indigo & Ice Blue)
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid') # Navy
    zebra_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid') # Very light gray/blue
    metric_fill = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid') # Soft blue
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Soft green
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Soft red
    
    title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=12, bold=True, color='1A365D')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Segoe UI', size=11, bold=True, color='000000')
    regular_font = Font(name='Segoe UI', size=11, color='333333')
    pass_font = Font(name='Segoe UI', size=11, bold=True, color='006100')
    fail_font = Font(name='Segoe UI', size=11, bold=True, color='9C0006')
    
    thin_border = Border(
        left=Side(style='thin', color='D2D6DC'),
        right=Side(style='thin', color='D2D6DC'),
        top=Side(style='thin', color='D2D6DC'),
        bottom=Side(style='thin', color='D2D6DC')
    )
    
    # Title Banner
    ws_dash.merge_cells('A1:D2')
    title_cell = ws_dash['A1']
    title_cell.value = "Load Testing Performance Dashboard"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Metadata Block
    ws_dash['A4'] = "Test Metadata"
    ws_dash['A4'].font = section_font
    
    metadata = [
        ("Target URL", TARGET_URL),
        ("Concurrency", f"{CONCURRENT_USERS} Virtual Users"),
        ("Duration", f"{DURATION_SECONDS} Seconds"),
        ("Start Time", start_ts.strftime("%Y-%m-%d %H:%M:%S")),
        ("End Time", end_ts.strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for idx, (label, val) in enumerate(metadata, 5):
        ws_dash.cell(row=idx, column=1, value=label).font = bold_font
        ws_dash.cell(row=idx, column=2, value=val).font = regular_font
        ws_dash.cell(row=idx, column=1).border = thin_border
        ws_dash.cell(row=idx, column=2).border = thin_border
        
    # Divider
    for col in range(1, 5):
        ws_dash.cell(row=11, column=col).border = Border(bottom=Side(style='medium', color='1A365D'))
        
    # Main Metrics Tables
    ws_dash['A13'] = "Core Performance Metrics"
    ws_dash['A13'].font = section_font
    
    metric_headers = ["Metric Description", "Value"]
    for col_idx, header in enumerate(metric_headers, 1):
        cell = ws_dash.cell(row=14, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
    metrics = [
        ("Total Requests Sent", total),
        ("Successful Requests (200 OK)", successes),
        ("Failed Requests", failures),
        ("Success Rate (%)", f"{success_rate:.2f}%"),
        ("Average Requests/Sec (RPS)", f"{rps:.2f} req/s")
    ]
    
    for idx, (label, val) in enumerate(metrics, 15):
        ws_dash.cell(row=idx, column=1, value=label).font = bold_font
        cell_val = ws_dash.cell(row=idx, column=2, value=val)
        cell_val.font = regular_font
        ws_dash.cell(row=idx, column=1).border = thin_border
        cell_val.border = thin_border
        
        # Color coding success and failures
        if label.startswith("Success Rate"):
            cell_val.fill = pass_fill
            cell_val.font = pass_font
        elif label.startswith("Failed"):
            cell_val.fill = fail_fill if failures > 0 else zebra_fill
            cell_val.font = fail_font if failures > 0 else regular_font
            
    # Latency Stats Table
    ws_dash['C13'] = "Response Latency Metrics"
    ws_dash['C13'].font = section_font
    
    for col_idx, header in enumerate(metric_headers, 3):
        cell = ws_dash.cell(row=14, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
        
    latency_metrics = [
        ("Minimum Latency", f"{min_lat:.2f} ms"),
        ("Median (50th Pct) Latency", f"{med_lat:.2f} ms"),
        ("Average Latency", f"{avg_lat:.2f} ms"),
        ("90th Percentile Latency", f"{p90:.2f} ms"),
        ("95th Percentile Latency", f"{p95:.2f} ms"),
        ("99th Percentile Latency", f"{p99:.2f} ms"),
        ("Maximum Latency", f"{max_lat:.2f} ms")
    ]
    
    for idx, (label, val) in enumerate(latency_metrics, 15):
        ws_dash.cell(row=idx, column=3, value=label).font = bold_font
        cell_val = ws_dash.cell(row=idx, column=4, value=val)
        cell_val.font = regular_font
        ws_dash.cell(row=idx, column=3).border = thin_border
        cell_val.border = thin_border
        cell_val.fill = metric_fill

    # Adjust columns widths on dashboard
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 30
    ws_dash.column_dimensions['D'].width = 20
    
    # -------------------------------------------------------------
    # SHEET 2: Detailed Request Log
    # -------------------------------------------------------------
    ws_log = wb.create_sheet(title="Detailed Request Log")
    ws_log.views.sheetView[0].showGridLines = True
    
    log_headers = ["Timestamp", "Virtual User ID", "HTTP Status Code", "Latency (ms)"]
    
    for col_idx, header in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
    # Write details
    print("Writing detailed logs to second tab...")
    for row_idx, r in enumerate(results, 2):
        cell_time = ws_log.cell(row=row_idx, column=1, value=r["timestamp"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
        cell_user = ws_log.cell(row=row_idx, column=2, value=r["user_id"])
        cell_status = ws_log.cell(row=row_idx, column=3, value=r["status_code"])
        cell_latency = ws_log.cell(row=row_idx, column=4, value=round(r["latency_ms"], 2))
        
        cell_time.alignment = Alignment(horizontal='center')
        cell_user.alignment = Alignment(horizontal='center')
        cell_status.alignment = Alignment(horizontal='center')
        cell_latency.alignment = Alignment(horizontal='right')
        
        cell_time.font = regular_font
        cell_user.font = regular_font
        cell_status.font = regular_font
        cell_latency.font = regular_font
        
        cell_time.border = thin_border
        cell_user.border = thin_border
        cell_status.border = thin_border
        cell_latency.border = thin_border
        
        # Color code response status
        if r["status_code"] == 200:
            cell_status.fill = pass_fill
            cell_status.font = pass_font
        else:
            cell_status.fill = fail_fill
            cell_status.font = fail_font
            
        # Optional Zebra striping
        if row_idx % 2 == 1:
            if r["status_code"] == 200:
                cell_time.fill = zebra_fill
                cell_user.fill = zebra_fill
                cell_latency.fill = zebra_fill

    # Autofit log column widths
    ws_log.column_dimensions['A'].width = 26
    ws_log.column_dimensions['B'].width = 18
    ws_log.column_dimensions['C'].width = 20
    ws_log.column_dimensions['D'].width = 18
    
    # Save Workbook
    report_path = "automated_test/load_test_results.xlsx"
    wb.save(report_path)
    print(f"SUCCESS: Load test report successfully created at {report_path}")
    print("=" * 70 + "\n")

if __name__ == "__main__":
    run_load_test()
